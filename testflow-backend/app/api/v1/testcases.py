from io import BytesIO
from collections import defaultdict

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.database import get_db
from app.core.deps import get_current_user
from app.schemas.common import ResponseModel
from app.schemas.testcase import TestCaseCreate, TestCaseUpdate, TestCaseOut
from app.crud import crud_testcase
from app.models.module import Module

router = APIRouter(prefix="/testcases", tags=["测试用例"])

# ── xlsx 样式常量 ──
HEADER_FONT = Font(name='微软雅黑', size=12, bold=True, color='FF000000')
HEADER_FILL = PatternFill(fgColor='BDD7EE', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center')
THIN_BORDER = Border(
    top=Side('thin'), bottom=Side('thin'),
    left=Side('thin'), right=Side('thin')
)
DATA_FONT = Font(name='等线', size=11)
DATA_ALIGN = Alignment(vertical='center', wrap_text=True)
HEADERS = ['测试用例ID', '模块', '标题', '前置条件', '测试数据', '测试步骤', '预期结果', '实际结果']
FIELD_MAP = ['case_no', 'module', 'title', 'preconditions', 'test_data', 'test_steps', 'expected_result', 'actual_result']
COL_WIDTHS = [20.75, 27.375, 29.125, 24.875, 24.25, 30.25, 34.125, 9.75]


def _apply_header_style(ws):
    """给 Sheet 第一行写表头并应用样式"""
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 18


def _apply_data_style(cell):
    """给数据单元格应用样式"""
    cell.font = DATA_FONT
    cell.alignment = DATA_ALIGN


def _to_out(case, db: Session) -> dict:
    return TestCaseOut(
        id=case.id, case_no=case.case_no, title=case.title,
        priority=case.priority, exec_status=case.exec_status,
        executor=crud_testcase.get_executor_name(db, case.executor_id),
        project=crud_testcase.get_project_name(db, case.project_id),
        project_id=case.project_id,
        sprint_id=case.sprint_id,
        sprint_name=crud_testcase.get_sprint_name(db, case.sprint_id),
        module_id=case.module_id,
        module=case.module,
        module_name=crud_testcase.get_module_name(db, case.module_id),
        case_type=case.case_type or "ui",
        automation_status=case.automation_status or "not_generated",
        automation_path=case.automation_path or "",
        selector_path=case.selector_path or "",
        source=case.source or "manual",
        version=case.version or "v1.0",
        fingerprint=case.fingerprint or "",
        raw_data=case.raw_data or {},
        coverage_count=crud_testcase.get_coverage_count(db, case.id),
        preconditions=case.preconditions or "",
        test_data=case.test_data or "",
        test_steps=case.test_steps or "",
        expected_result=case.expected_result or "",
        actual_result=case.actual_result or "",
        updated_at=case.updated_at,
    ).model_dump()


@router.get("", response_model=ResponseModel)
def list_testcases(
    project: str | None = None,
    keyword: str | None = Query(None, description="搜索关键词"),
    sprint_id: int | None = Query(None, description="Sprint ID"),
    module_id: int | None = Query(None, description="模块 ID"),
    case_type: str | None = Query(None, description="用例类型"),
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    cases = crud_testcase.get_testcases(
        db,
        project=project,
        keyword=keyword,
        sprint_id=sprint_id,
        module_id=module_id,
        case_type=case_type,
    )
    return ResponseModel(data=[_to_out(c, db) for c in cases])


@router.get("/stats", response_model=ResponseModel)
def testcase_stats(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return ResponseModel(data=crud_testcase.get_testcase_stats(db))


@router.post("/batch-execute", response_model=ResponseModel)
def batch_execute(
    project_id: int | None = None,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user)
):
    count = crud_testcase.batch_execute(db, project_id)
    return ResponseModel(data={"executed_count": count}, message=f"已批量执行 {count} 条用例")


@router.get("/export")
def export_testcases(
    project_id: int | None = Query(None, description="项目ID，不传则导出全部"),
    sprint_id: int | None = Query(None, description="Sprint ID"),
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    """导出测试用例为 xlsx，按模块分 Sheet，模块列显示中文名"""
    # 获取用例
    query = db.query(crud_testcase.TestCase).filter(crud_testcase.TestCase.is_deleted == False)  # noqa: E712
    if project_id:
        query = query.filter(crud_testcase.TestCase.project_id == project_id)
    if sprint_id:
        query = query.filter(crud_testcase.TestCase.sprint_id == sprint_id)
    cases = query.all()

    if not cases:
        raise HTTPException(status_code=404, detail="没有可导出的测试用例")

    # 预加载 module_id → Module 映射
    module_map = {}
    for c in cases:
        if c.module_id and c.module_id not in module_map:
            mod = db.query(Module).filter(Module.id == c.module_id).first()
            if mod:
                module_map[c.module_id] = mod

    # 按模块中文名分组
    grouped = defaultdict(list)
    for c in cases:
        mod = module_map.get(c.module_id) if c.module_id else None
        display_name = mod.name if mod else (c.module or '未分类')
        grouped[display_name].append(c)

    wb = openpyxl.Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    for module_name, module_cases in grouped.items():
        # Sheet 名最长 31 字符
        sheet_title = module_name[:31]
        ws = wb.create_sheet(title=sheet_title)
        _apply_header_style(ws)

        for row_idx, case in enumerate(module_cases, start=2):
            # 模块列：优先显示中文名
            mod = module_map.get(case.module_id) if case.module_id else None
            display_module = mod.name if mod else (case.module or '')
            values = [
                case.case_no or '', display_module, case.title or '',
                case.preconditions or '', case.test_data or '',
                case.test_steps or '', case.expected_result or '',
                case.actual_result or ''
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                _apply_data_style(cell)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=testcases.xlsx'}
    )


@router.get("/template")
def download_template(_=Depends(get_current_user)):
    """下载测试用例导入模板（只有表头的空 xlsx）"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '测试用例'
    _apply_header_style(ws)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename=testcase_template.xlsx'}
    )


@router.post("/import", response_model=ResponseModel)
async def import_testcases(
    file: UploadFile = File(...),
    project_id: int = Form(...),
    sprint_id: int | None = Form(None),
    db: Session = Depends(get_db),
    _=Depends(get_current_user)
):
    """导入测试用例 xlsx 文件"""
    # 校验项目存在
    from app.models.project import Project
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=400, detail="项目不存在")

    # 校验文件类型
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    # 读取文件
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="文件解析失败，请检查文件格式")

    rows = []
    # Excel 列名 → 字段名 映射
    col_map = {h: FIELD_MAP[i] for i, h in enumerate(HEADERS)}

    for ws in wb.worksheets:
        # 读取表头，建立列索引
        header_cells = [cell.value for cell in ws[1]]
        col_indexes = {}
        for col_idx, header_val in enumerate(header_cells):
            if header_val and header_val in col_map:
                col_indexes[col_map[header_val]] = col_idx  # 0-based index

        if not col_indexes:
            continue  # 跳过无有效表头的 Sheet

        for row in ws.iter_rows(min_row=2, values_only=True):
            # 跳过全空行
            if not any(v for v in row):
                continue
            row_data = {}
            for field_name, col_idx in col_indexes.items():
                if col_idx < len(row):
                    row_data[field_name] = str(row[col_idx] or '')
            rows.append(row_data)
            if sprint_id:
                row_data['sprint_id'] = sprint_id

    if not rows:
        raise HTTPException(status_code=400, detail="文件中没有有效数据")

    result = crud_testcase.import_testcases(db, project_id, rows)
    parts = [f"新增 {result['success_count']} 条"]
    if result.get('updated_count', 0) > 0:
        parts.append(f"更新 {result['updated_count']} 条")
    if result['fail_count'] > 0:
        parts.append(f"失败 {result['fail_count']} 条")
    return ResponseModel(data=result, message=f"导入完成：{'，'.join(parts)}")


@router.post("", response_model=ResponseModel)
def create_testcase(data: TestCaseCreate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    case = crud_testcase.create_testcase(db, data)
    return ResponseModel(data=_to_out(case, db))


@router.put("/{case_id}", response_model=ResponseModel)
def update_testcase(case_id: int, data: TestCaseUpdate, db: Session = Depends(get_db), _=Depends(get_current_user)):
    case = crud_testcase.get_testcase(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="用例不存在")
    case = crud_testcase.update_testcase(db, case, data)
    return ResponseModel(data=_to_out(case, db))


@router.delete("/{case_id}", response_model=ResponseModel)
def delete_testcase(case_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    case = crud_testcase.get_testcase(db, case_id)
    if not case:
        raise HTTPException(status_code=404, detail="用例不存在")
    crud_testcase.delete_testcase(db, case)
    return ResponseModel(message="删除成功")
