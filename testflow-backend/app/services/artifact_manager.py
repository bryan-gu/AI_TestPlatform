"""
SKILL 产物文件管理器

负责管理流水线生成的所有文件产物（功能点.md、cases.json、Excel 等），
存为 Document 记录挂到对应 Sprint 知识库，前端可见可下载。

磁盘存储结构：
  uploads/artifacts/{sprint_id}/
      ├── features/{module_name}/功能点.md
      ├── cases/{module_name}/cases.json
      └── testcases/{sprint_name}_testCase.xlsx
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime

from sqlalchemy.orm import Session

from app.models.document import Document
from app.models.module import Module
from app.models.sprint import Sprint
from app.crud import crud_knowledge_asset

logger = logging.getLogger(__name__)

# artifacts 根目录（相对于项目根目录）
ARTIFACTS_ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "uploads", "artifacts")


class ArtifactManager:
    """SKILL 产物文件管理"""

    def __init__(self, db: Session, sprint_id: int):
        self.db = db
        self.sprint_id = sprint_id
        self.sprint = db.query(Sprint).filter(Sprint.id == sprint_id).first()
        self.sprint_name = self.sprint.name if self.sprint else f"sprint{sprint_id}"
        self.base_dir = os.path.join(ARTIFACTS_ROOT, str(sprint_id))
        os.makedirs(self.base_dir, exist_ok=True)

    # ============ 功能点.md ============

    def save_feature_points_md(self, module_name: str, content: str, source_doc_id: int | None = None) -> Document:
        """
        保存功能点.md 文件，同时创建 Document 记录

        Args:
            module_name: 模块名称
            content: Markdown 内容
            source_doc_id: 来源文档 ID

        Returns:
            Document 记录
        """
        # 写磁盘文件
        rel_path = f"features/{module_name}"
        abs_dir = os.path.join(self.base_dir, rel_path)
        os.makedirs(abs_dir, exist_ok=True)
        file_path = os.path.join(abs_dir, "功能点.md")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)

        # 创建/更新 Document 记录
        doc_name = f"{module_name}-功能点"
        doc = self._upsert_document(
            name=doc_name,
            file_path=file_path,
            file_type="Markdown",
            content_preview=content[:500],
            ai_summary=f"AI 提取的 {module_name} 模块功能点",
            ai_status="AI生成",
        )
        crud_knowledge_asset.upsert_asset_for_document(
            self.db,
            doc,
            project_id=self.sprint.project_id if self.sprint else None,
            source_kind="ai_generated",
            asset_type="feature_spec",
            module_id=self._get_module_id(module_name),
            metadata={"artifact_kind": "feature_points_md", "module_name": module_name},
        )
        return doc

    # ============ cases.json ============

    def save_cases_json(self, module_name: str, cases: list[dict]) -> Document:
        """
        保存 cases.json 文件，同时创建 Document 记录

        Args:
            module_name: 模块名称
            cases: 测试用例列表（JSON 格式）

        Returns:
            Document 记录
        """
        content = json.dumps(cases, ensure_ascii=False, indent=2)

        # 写磁盘文件
        rel_path = f"cases/{module_name}"
        abs_dir = os.path.join(self.base_dir, rel_path)
        os.makedirs(abs_dir, exist_ok=True)
        file_path = os.path.join(abs_dir, "cases.json")
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(content)

        # 创建/更新 Document 记录
        doc_name = f"{module_name}-测试用例"
        doc = self._upsert_document(
            name=doc_name,
            file_path=file_path,
            file_type="JSON",
            content_preview=f"共 {len(cases)} 条测试用例",
            ai_summary=f"{module_name} 模块的 {len(cases)} 条结构化测试用例",
            ai_status="AI生成",
        )
        crud_knowledge_asset.upsert_asset_for_document(
            self.db,
            doc,
            project_id=self.sprint.project_id if self.sprint else None,
            source_kind="ai_generated",
            asset_type="test_case_json",
            module_id=self._get_module_id(module_name),
            metadata={"artifact_kind": "cases_json", "module_name": module_name, "case_count": len(cases)},
        )
        return doc

    # ============ testCase.xlsx ============

    def save_excel(self, all_cases: dict[str, list[dict]]) -> Document | None:
        """
        生成并保存 testCase.xlsx（跨模块汇总）

        Args:
            all_cases: {模块名: [用例列表]}

        Returns:
            Document 记录，如果没有用例则返回 None
        """
        total = sum(len(cases) for cases in all_cases.values())
        if total == 0:
            return None

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.warning("openpyxl 未安装，跳过 Excel 生成")
            return None

        # 样式
        header_font = Font(name="微软雅黑", size=12, bold=True)
        header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
        header_align = Alignment(wrap_text=True, vertical="center")
        data_font = Font(name="微软雅黑", size=11)
        data_align = Alignment(wrap_text=True, vertical="top")
        col_widths = [14.875, 16.625, 20.375, 21.125, 17.875, 26.5, 25.875, 8.675]
        headers = ["测试用例ID", "模块", "标题", "前置条件", "测试数据", "测试步骤", "预期结果", "实际结果"]

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for module_name, cases in all_cases.items():
            sheet_name = module_name[:31]
            ws = wb.create_sheet(title=sheet_name)

            # 表头
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = header_font
                c.fill = header_fill
                c.alignment = header_align
            for ci, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

            # 数据行
            keys = ["id", "module", "title", "precondition", "test_data", "steps", "expected"]
            for row_idx, case in enumerate(cases, 2):
                for ci, key in enumerate(keys, 1):
                    c = ws.cell(row=row_idx, column=ci, value=case.get(key, ""))
                    c.font = data_font
                    c.alignment = data_align
                ws.cell(row=row_idx, column=8, value="")

        # 保存文件
        abs_dir = os.path.join(self.base_dir, "testcases")
        os.makedirs(abs_dir, exist_ok=True)
        file_name = f"{self.sprint_name}_testCase.xlsx"
        file_path = os.path.join(abs_dir, file_name)
        wb.save(file_path)

        # 创建 Document 记录
        doc_name = f"{self.sprint_name}-测试用例汇总"
        doc = self._upsert_document(
            name=doc_name,
            file_path=file_path,
            file_type="Excel",
            content_preview=f"共 {total} 条测试用例，覆盖 {len(all_cases)} 个模块",
            ai_summary=f"{self.sprint_name} Sprint 测试用例汇总 Excel，共 {total} 条",
            ai_status="AI生成",
        )
        crud_knowledge_asset.upsert_asset_for_document(
            self.db,
            doc,
            project_id=self.sprint.project_id if self.sprint else None,
            source_kind="ai_generated",
            asset_type="test_case_excel",
            metadata={"artifact_kind": "testcase_excel", "case_count": total, "module_count": len(all_cases)},
        )
        return doc

    # ============ 读取文件 ============

    def read_cases_json(self, module_name: str) -> list[dict]:
        """读取已保存的 cases.json"""
        file_path = os.path.join(self.base_dir, "cases", module_name, "cases.json")
        if not os.path.exists(file_path):
            return []
        with open(file_path, "r", encoding="utf-8") as f:
            return json.load(f)

    def read_feature_points_md(self, module_name: str) -> str:
        """读取已保存的功能点.md"""
        file_path = os.path.join(self.base_dir, "features", module_name, "功能点.md")
        if not os.path.exists(file_path):
            return ""
        with open(file_path, "r", encoding="utf-8") as f:
            return f.read()

    # ============ 内部方法 ============

    def _get_module_id(self, module_name: str) -> int | None:
        if not module_name:
            return None
        query = self.db.query(Module).filter(Module.name == module_name)
        if self.sprint and self.sprint.project_id:
            query = query.filter(Module.project_id == self.sprint.project_id)
        module = query.first()
        return module.id if module else None

    def _upsert_document(
        self,
        name: str,
        file_path: str,
        file_type: str,
        content_preview: str = "",
        ai_summary: str = "",
        ai_status: str = "AI生成",
    ) -> Document:
        """
        创建或更新 Document 记录。
        同名 + 同 Sprint + AI生成的文档只保留一份（更新内容）。
        """
        existing = self.db.query(Document).filter(
            Document.sprint_id == self.sprint_id,
            Document.name == name,
            Document.ai_status == "AI生成",
        ).first()

        if existing:
            # 更新
            existing.file_path = file_path
            existing.file_type = file_type
            existing.file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
            existing.content_preview = content_preview
            existing.ai_summary = ai_summary
            existing.version = "v1.0"
            self.db.commit()
            self.db.refresh(existing)
            return existing
        else:
            # 新建
            doc = Document(
                name=name,
                file_path=file_path,
                file_type=file_type,
                file_size=os.path.getsize(file_path) if os.path.exists(file_path) else 0,
                sprint_id=self.sprint_id,
                uploader_id=None,  # AI 生成，无上传者
                content_preview=content_preview,
                ai_summary=ai_summary,
                ai_status=ai_status,
            )
            self.db.add(doc)
            self.db.commit()
            self.db.refresh(doc)
            return doc

    def get_excel_path(self) -> str | None:
        """获取 Excel 文件路径（用于下载）"""
        abs_dir = os.path.join(self.base_dir, "testcases")
        if not os.path.exists(abs_dir):
            return None
        for f in os.listdir(abs_dir):
            if f.endswith(".xlsx"):
                return os.path.join(abs_dir, f)
        return None
