"""本地项目资料导入服务（阶段 7）。

扫描本地目录树，按约定结构识别 Sprint 与知识资产，支持 dry_run 预览与正式导入。
正式导入会创建/更新 Document、KnowledgeAsset，结构化导入 Excel 用例、cases.json 用例、
OpenAPI / Markdown 接口端点，并回填 TraceLink、重建图谱。

约定目录结构（识别规则，大小写不敏感）：
  需求文档/sprintN/*.md|*.pdf|*.docx        -> requirement_doc
  需求文档/sprintN/.../功能点.md              -> feature_spec
  需求文档/sprintN/.../cases.json             -> test_case_json
  测试用例/sprintN/*testCase.xlsx             -> test_case_excel
  测试用例/sprintN/*_scripts/*.spec.ts        -> test_script
  测试用例/sprintN/*_scripts/selectors/*.selectors.ts -> selector_map
  接口文档/*.md                                -> api_doc_md
  接口文档/*.json (OpenAPI 合法)               -> api_doc_openapi
未归属某个 sprintN 的资产归入项目 sprint_all 最新汇总。
"""
import os
import re
import json
import logging
from pathlib import Path

from sqlalchemy.orm import Session

from app.crud import crud_sprint, crud_knowledge_asset, crud_trace_link, crud_graph, crud_testcase
from app.models.document import Document
from app.models.sprint import Sprint
from app.schemas.knowledge_asset import (
    LocalProjectImportResult,
    LocalProjectImportAssetPreview,
    LocalProjectImportSprintPreview,
)
from app.schemas.trace_link import TraceLinkCreate
from app.services.api_doc_parser import parse_openapi_file, parse_markdown_file
from app.services.api_import_service import upsert_endpoints_for_asset

logger = logging.getLogger(__name__)

SKIP_DIRS = {
    ".git", "node_modules", "venv", ".venv", "dist", "build",
    "__pycache__", ".idea", ".vscode", "target", ".pytest_cache",
}
MAX_FILE_SIZE = 50 * 1024 * 1024
MAX_TOTAL_FILES = 2000

SPRINT_SEG_RE = re.compile(r"^sprint[\-_ ]?([0-9a-zA-Z]+)$", re.IGNORECASE)

# 与 testcases.py 导入保持一致的 Excel 表头
EXCEL_HEADERS = ['测试用例ID', '模块', '标题', '前置条件', '测试数据', '测试步骤', '预期结果', '实际结果']
EXCEL_FIELD_MAP = ['case_no', 'module', 'title', 'preconditions', 'test_data', 'test_steps', 'expected_result', 'actual_result']


class LocalProjectImporter:
    def __init__(self, db: Session):
        self.db = db

    # ============ 扫描 ============

    def scan(self, root_path: str, project_id: int) -> LocalProjectImportResult:
        root = self._validate_root(root_path)
        assets: list[LocalProjectImportAssetPreview] = []
        warnings: list[str] = []
        sprint_counts: dict[str, int] = {}

        for abs_path, rel_path, name in self._walk(root):
            asset_type, file_type, file_size = self._classify(rel_path, name, abs_path)
            sprint_name = self._detect_sprint(rel_path)
            display_sprint = sprint_name or "sprint_all"
            sprint_counts[display_sprint] = sprint_counts.get(display_sprint, 0) + 1
            assets.append(LocalProjectImportAssetPreview(
                rel_path=rel_path,
                name=name,
                asset_type=asset_type,
                file_type=file_type,
                file_size=file_size,
                sprint_name=display_sprint,
                module_hint=self._module_hint(rel_path),
            ))

        counts = self._summarize_asset_types(assets)
        if not assets:
            warnings.append("未在目录中识别到任何可导入的资料文件")

        return LocalProjectImportResult(
            root_path=str(root),
            project_id=project_id,
            dry_run=True,
            sprints=[LocalProjectImportSprintPreview(name=k, asset_count=v)
                     for k, v in sorted(sprint_counts.items())],
            assets=assets,
            counts=counts,
            warnings=warnings,
        )

    # ============ 正式导入 ============

    def import_project(self, root_path: str, project_id: int, dry_run: bool = False) -> LocalProjectImportResult:
        preview = self.scan(root_path, project_id)
        if dry_run:
            return preview

        touched_sprint_ids: set[int] = set()
        imported = {
            "documents": 0, "assets": 0, "testcases": 0,
            "api_endpoints": 0, "trace_links": 0,
        }
        warnings = list(preview.warnings)

        for asset_pv in preview.assets:
            try:
                sprint = self._ensure_sprint(project_id, asset_pv.sprint_name)
                touched_sprint_ids.add(sprint.id)
                abs_path = os.path.join(root_path, asset_pv.rel_path)
                doc = self._upsert_document(
                    sprint, asset_pv.name, abs_path,
                    asset_pv.file_type, asset_pv.file_size,
                )
                imported["documents"] += 1
                ka = crud_knowledge_asset.upsert_asset_for_document(
                    self.db, doc,
                    project_id=project_id,
                    source_kind="imported",
                    asset_type=asset_pv.asset_type,
                    metadata={"local_import": True, "rel_path": asset_pv.rel_path},
                )
                imported["assets"] += 1
                # asset -> document / derived_from
                crud_trace_link.upsert_trace_link(self.db, TraceLinkCreate(
                    project_id=project_id, sprint_id=sprint.id,
                    source_type="asset", source_id=ka.id,
                    target_type="document", target_id=doc.id,
                    relation_type="derived_from", confidence=100,
                    evidence="本地导入资产关联文档",
                    metadata={"import_type": "local_project"},
                    created_by="local-import",
                ), commit=False)
                imported["trace_links"] += 1

                # 结构化导入
                self._structured_import(ka, doc, sprint, project_id, abs_path,
                                        asset_pv.asset_type, imported, warnings)
            except Exception as e:
                logger.exception("本地导入文件失败: %s", asset_pv.rel_path)
                warnings.append(f"导入失败 {asset_pv.rel_path}: {e}")

        self.db.commit()

        # 回填 TraceLink + 重建图谱
        try:
            backfill = crud_trace_link.backfill_trace_links(self.db, project_id=project_id)
            imported["trace_links"] += backfill.get("upserted_count", 0)
        except Exception as e:
            warnings.append(f"TraceLink 回填失败: {e}")

        graph_info = {}
        for sid in touched_sprint_ids:
            try:
                g = crud_graph.generate_graph_for_scope(self.db, project_id=project_id, sprint_id=sid)
                if not graph_info:
                    graph_info = {
                        "id": g.id, "node_count": g.node_count or 0,
                        "edge_count": g.edge_count or 0, "sprint_id": sid,
                    }
            except Exception as e:
                warnings.append(f"图谱重建失败 sprint_id={sid}: {e}")

        preview.dry_run = False
        preview.imported = imported
        preview.graph = graph_info
        preview.warnings = warnings
        return preview

    # ============ 结构化导入 ============

    def _structured_import(self, asset, doc, sprint, project_id, abs_path, asset_type, imported, warnings):
        if asset_type == "test_case_excel":
            rows = self._parse_excel_rows(abs_path)
            for r in rows:
                r["sprint_id"] = sprint.id
            if rows:
                res = crud_testcase.import_testcases(self.db, project_id, rows)
                imported["testcases"] += res.get("success_count", 0) + res.get("updated_count", 0)
            else:
                warnings.append(f"Excel 无有效数据：{asset.name}")

        elif asset_type == "test_case_json":
            rows = self._parse_cases_json(abs_path)
            for r in rows:
                r["sprint_id"] = sprint.id
            if rows:
                res = crud_testcase.import_testcases(self.db, project_id, rows)
                imported["testcases"] += res.get("success_count", 0) + res.get("updated_count", 0)
            else:
                warnings.append(f"cases.json 无有效用例：{asset.name}")

        elif asset_type == "api_doc_openapi":
            try:
                endpoints = parse_openapi_file(abs_path)
                if endpoints:
                    created, updated = upsert_endpoints_for_asset(self.db, asset, endpoints)
                    imported["api_endpoints"] += created + updated
                else:
                    warnings.append(f"OpenAPI 未解析出接口：{asset.name}")
            except Exception as e:
                warnings.append(f"OpenAPI 解析失败 {asset.name}: {e}")

        elif asset_type == "api_doc_md":
            try:
                endpoints, md_warnings = parse_markdown_file(abs_path)
                warnings.extend(md_warnings)
                if endpoints:
                    created, updated = upsert_endpoints_for_asset(self.db, asset, endpoints)
                    imported["api_endpoints"] += created + updated
            except Exception as e:
                warnings.append(f"Markdown 接口解析失败 {asset.name}: {e}")

    def _parse_excel_rows(self, abs_path: str) -> list[dict]:
        try:
            import openpyxl
            from io import BytesIO
        except ImportError:
            return []
        try:
            wb = openpyxl.load_workbook(abs_path, read_only=True, data_only=True)
        except Exception:
            return []
        col_map = {h: EXCEL_FIELD_MAP[i] for i, h in enumerate(EXCEL_HEADERS)}
        rows: list[dict] = []
        for ws in wb.worksheets:
            header_cells = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            col_indexes = {}
            for idx, hv in enumerate(header_cells):
                if hv and hv in col_map:
                    col_indexes[col_map[hv]] = idx
            if not col_indexes:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(v for v in row):
                    continue
                rd = {}
                for field, idx in col_indexes.items():
                    if idx < len(row):
                        rd[field] = str(row[idx] or "")
                rows.append(rd)
        return rows

    def _parse_cases_json(self, abs_path: str) -> list[dict]:
        try:
            with open(abs_path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return []
        if isinstance(data, dict):
            for key in ("test_cases", "cases", "testCases", "data", "items"):
                if isinstance(data.get(key), list):
                    data = data[key]
                    break
        if not isinstance(data, list):
            return []
        rows = []
        for item in data:
            if not isinstance(item, dict):
                continue
            title = (item.get("title") or "").strip()
            if not title:
                continue
            rows.append({
                "case_no": item.get("id") or item.get("case_no") or "",
                "module": item.get("module") or "",
                "title": title,
                "preconditions": item.get("precondition") or item.get("preconditions") or "",
                "test_data": item.get("test_data") or "",
                "test_steps": item.get("steps") or item.get("test_steps") or "",
                "expected_result": item.get("expected") or item.get("expected_result") or "",
            })
        return rows

    # ============ 文档/Sprint upsert ============

    def _upsert_document(self, sprint: Sprint, name: str, abs_path: str,
                         file_type: str, file_size: int) -> Document:
        existing = self.db.query(Document).filter(
            Document.sprint_id == sprint.id,
            Document.name == name,
            Document.file_path == abs_path,
            Document.is_deleted == False,  # noqa: E712
        ).first()
        if existing:
            existing.file_type = file_type
            existing.file_size = file_size or (os.path.getsize(abs_path) if os.path.exists(abs_path) else 0)
            self.db.commit()
            self.db.refresh(existing)
            return existing
        doc = Document(
            name=name,
            file_path=abs_path,
            file_type=file_type,
            file_size=file_size or (os.path.getsize(abs_path) if os.path.exists(abs_path) else 0),
            sprint_id=sprint.id,
            uploader_id=None,
            ai_status="待分析",
            parse_status="待解析",
        )
        self.db.add(doc)
        self.db.commit()
        self.db.refresh(doc)
        return doc

    def _ensure_sprint(self, project_id: int, sprint_name: str) -> Sprint:
        if not sprint_name or sprint_name == "sprint_all":
            return crud_sprint.ensure_sprint_all(self.db, project_id)
        norm = crud_sprint.normalize_sprint_name(sprint_name)
        existing = self.db.query(Sprint).filter(
            Sprint.project_id == project_id,
            Sprint.is_deleted == False,  # noqa: E712
            Sprint.is_all == False,  # noqa: E712
        ).all()
        for s in existing:
            if crud_sprint.normalize_sprint_name(s.name) == norm:
                return s
        sprint = Sprint(name=sprint_name, project_id=project_id, status="待启动", is_all=False)
        self.db.add(sprint)
        self.db.commit()
        self.db.refresh(sprint)
        return sprint

    # ============ 路径解析与分类 ============

    def _validate_root(self, root_path: str) -> Path:
        if not root_path or not root_path.strip():
            raise ValueError("root_path 不能为空")
        root = Path(root_path).expanduser().resolve()
        if not root.exists():
            raise ValueError(f"路径不存在: {root}")
        if not root.is_dir():
            raise ValueError(f"路径不是目录: {root}")
        return root

    def _walk(self, root: Path):
        count = 0
        for dirpath, dirnames, filenames in os.walk(root):
            dirnames[:] = [d for d in dirnames if d not in SKIP_DIRS and not d.startswith(".")]
            for fn in filenames:
                if fn.startswith("."):
                    continue
                abs_path = os.path.join(dirpath, fn)
                if not os.path.isfile(abs_path):
                    continue
                if os.path.getsize(abs_path) > MAX_FILE_SIZE:
                    continue
                rel_path = os.path.relpath(abs_path, root)
                yield abs_path, rel_path, fn
                count += 1
                if count >= MAX_TOTAL_FILES:
                    return

    def _classify(self, rel_path: str, name: str, abs_path: str) -> tuple[str, str, int]:
        ext = os.path.splitext(name)[1].lower()
        lower_rel = rel_path.replace("\\", "/").lower()
        file_size = os.path.getsize(abs_path) if os.path.exists(abs_path) else 0

        # 接口文档
        if "接口文档" in rel_path or "接口文档" in name or "api" in lower_rel or "接口" in name:
            if ext == ".json" and self._looks_openapi(abs_path):
                return "api_doc_openapi", self._ext_file_type(ext), file_size
            if ext == ".md":
                return "api_doc_md", self._ext_file_type(ext), file_size

        # 需求文档
        if "需求文档" in rel_path or "需求" in lower_rel:
            if "cases.json" in lower_rel or name.lower() == "cases.json":
                return "test_case_json", "JSON", file_size
            if "功能点" in rel_path or "功能点" in name or "feature" in lower_rel:
                if ext in (".md", ".pdf", ".docx", ".doc"):
                    return "feature_spec", self._ext_file_type(ext), file_size
            if ext in (".md", ".pdf", ".docx", ".doc"):
                return "requirement_doc", self._ext_file_type(ext), file_size

        # 测试用例
        if "测试用例" in rel_path or "testCase" in name or "testcase" in lower_rel:
            if ext == ".spec.ts" or name.endswith(".spec.ts"):
                return "test_script", "TypeScript", file_size
            if ext == ".selectors.ts" or name.endswith(".selectors.ts"):
                return "selector_map", "TypeScript", file_size
            if ext in (".xlsx", ".xls"):
                return "test_case_excel", "Excel", file_size

        # 兜底：按扩展名推断
        asset_type = crud_knowledge_asset.infer_asset_type(name, self._ext_file_type(ext), source_kind="imported")
        return asset_type, self._ext_file_type(ext), file_size

    def _detect_sprint(self, rel_path: str) -> str | None:
        for part in rel_path.replace("\\", "/").split("/"):
            m = SPRINT_SEG_RE.match(part)
            if m:
                return part
        return None

    def _looks_openapi(self, abs_path: str) -> bool:
        try:
            with open(abs_path, "r", encoding="utf-8") as f:
                head = f.read(2048)
        except Exception:
            return False
        return ("openapi" in head) or ("swagger" in head) or ('"paths"' in head)

    def _module_hint(self, rel_path: str) -> str:
        parts = rel_path.replace("\\", "/").split("/")
        for idx, p in enumerate(parts):
            if p in ("需求功能点", "功能点") and idx + 1 < len(parts):
                return parts[idx + 1]
        return ""

    def _ext_file_type(self, ext: str) -> str:
        return {
            ".md": "Markdown", ".pdf": "PDF", ".doc": "Word", ".docx": "Word",
            ".xlsx": "Excel", ".xls": "Excel", ".json": "JSON",
            ".ts": "TypeScript", ".js": "JavaScript",
            ".png": "PNG", ".jpg": "JPG", ".jpeg": "JPEG",
        }.get(ext, ext.lstrip(".").capitalize() or "其他")

    def _summarize_asset_types(self, assets: list[LocalProjectImportAssetPreview]) -> dict[str, int]:
        counts: dict[str, int] = {}
        for a in assets:
            counts[a.asset_type] = counts.get(a.asset_type, 0) + 1
        return counts
